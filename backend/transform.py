"""
LCM Import Tool — Transform Script
Writes data directly into the template's XML zip (bypassing openpyxl's save)
so conditional formatting, dropdowns, and cell colors are fully preserved.
"""
import csv, re, json, sys, zipfile, os
import openpyxl
from xml.sax.saxutils import escape as xml_escape

SKIP_FIELDS = {
    'mirakl-acceptance-status','mirakl-authorized-selling-shop-ids','mirakl-catalogs',
    'mirakl-creation-date','mirakl-integration-errors',
    'mirakl-last-operator-acceptance-action-date',
    'mirakl-last-operator-acceptance-action-user-name','mirakl-product-id',
    'mirakl-product-urls','mirakl-rejection-message','mirakl-rejection-reason',
    'mirakl-restricted-selling','mirakl-sources','mirakl-synchronization-status',
    'mirakl-update-date','mirakl-validation-status','vgc',
}
DEFAULT_TAX = 'P0000000'
SENTINEL = object()

def fix_encoding(t):
    if not t: return t
    try: return t.encode('cp1252').decode('utf-8')
    except: return t

def clean_value(field, value):
    if not value: return value
    value = fix_encoding(str(value)).replace('\xc4\xa2', '•')
    if field == 'itemRefundable': value = re.sub(r'^itemRefundable_', '', value)
    elif field == 'isBackordered': value = re.sub(r'^Is_Backordered_', '', value)
    return value

def to_number(val):
    if val is None or str(val).strip() == '': return None
    try:
        f = float(val)
        return int(f) if f == int(f) else f
    except: return None

def strip_prefix(sku): return re.sub(r'^SHOP\d+_SKU', '', sku)

def index_to_col_letter(idx):
    result = ''
    n = idx + 1
    while n:
        n, r = divmod(n - 1, 26)
        result = chr(65 + r) + result
    return result

def make_cell_xml(col_letter, row_num, value):
    ref = f"{col_letter}{row_num}"
    if isinstance(value, (int, float)):
        return f'<c r="{ref}"><v>{value}</v></c>'
    return f'<c r="{ref}" t="inlineStr"><is><t>{xml_escape(str(value))}</t></is></c>'

def transform(export_csv, offer_xlsx, template_xlsx, output_xlsx):
    # Load offers
    wb_o = openpyxl.load_workbook(offer_xlsx)
    ws_o = wb_o.active
    oh = [c.value for c in ws_o[1]]
    offers = {}
    for row in ws_o.iter_rows(min_row=2, values_only=True):
        o = dict(zip(oh, row))
        if o.get('Offer SKU'): offers[str(o['Offer SKU']).strip()] = o

    # Load products
    with open(export_csv, encoding='utf-8') as f:
        export_rows = list(csv.DictReader(f))

    # Get field→column mapping from template row 2
    wb_t = openpyxl.load_workbook(template_xlsx)
    ws_t = wb_t['Data']
    field_to_col = {str(c.value).strip(): c.column - 1 for c in ws_t[2] if c.value}
    wb_t.close()

    # Mismatch warnings
    product_base_skus = {strip_prefix(p['mirakl-product-sku'].strip()) for p in export_rows}
    unmatched_products = sorted(product_base_skus - set(offers.keys()))
    unmatched_offers   = sorted(set(offers.keys()) - product_base_skus)

    # Build row XML for each product
    row_xml_blocks = []
    for idx, exp in enumerate(export_rows):
        row_num  = idx + 3
        base_sku = strip_prefix(exp.get('mirakl-product-sku', '').strip())
        t_sku    = base_sku + '-Deals'
        offer    = offers.get(base_sku, {})
        cells    = {}

        for field, col_idx in field_to_col.items():
            val = SENTINEL
            if field in ('shopSku','upc','sku','product-id'): val = t_sku
            elif field == 'product-id-type': val = 'SHOP_SKU'
            elif field in ('vgc','vg-name'):
                v = clean_value('vg-name', exp.get('vg-name','').strip())
                val = v if v else SENTINEL
            elif field == 'state': val = 'New'
            elif field == 'price':
                n = to_number(offer.get('Original price')) if offer else None
                val = n if n is not None else SENTINEL
            elif field == 'quantity':
                if offer:
                    n = to_number(offer.get('Quantity'))
                    val = n if n is not None else SENTINEL
            elif field == 'discount-price':
                n = to_number(offer.get('Price')) if offer else None
                val = n if n is not None else SENTINEL
            elif field == 'product-tax-code':
                if offer:
                    t = str(offer.get('Product tax code') or '').strip()
                    val = t if t else DEFAULT_TAX
            elif field in exp and field not in SKIP_FIELDS:
                c = clean_value(field, exp[field].strip())
                val = c if c else SENTINEL
            if val is not SENTINEL: cells[col_idx] = val

        cell_xml = ''.join(
            make_cell_xml(index_to_col_letter(ci), row_num, v)
            for ci, v in sorted(cells.items())
        )
        row_xml_blocks.append(f'<row r="{row_num}">{cell_xml}</row>')

    # Read template zip, inject rows into sheet1.xml, write new zip
    with zipfile.ZipFile(template_xlsx, 'r') as zin:
        file_contents = {name: zin.read(name) for name in zin.namelist()}
        zip_infos     = {name: zin.getinfo(name) for name in zin.namelist()}

    sheet_xml = file_contents['xl/worksheets/sheet1.xml'].decode('utf-8')
    sheet_xml = sheet_xml.replace('</sheetData>', ''.join(row_xml_blocks) + '</sheetData>', 1)
    file_contents['xl/worksheets/sheet1.xml'] = sheet_xml.encode('utf-8')

    if os.path.exists(output_xlsx):
        os.chmod(output_xlsx, 0o644)
        os.remove(output_xlsx)

    with zipfile.ZipFile(output_xlsx, 'w', zipfile.ZIP_DEFLATED) as zout:
        for name, content in file_contents.items():
            zout.writestr(zip_infos[name], content)

    return {
        'products': len(export_rows),
        'offers_matched': sum(1 for p in export_rows if strip_prefix(p['mirakl-product-sku'].strip()) in offers),
        'unmatched_products': unmatched_products,
        'unmatched_offers': unmatched_offers,
    }

if __name__ == '__main__':
    if len(sys.argv) != 5:
        print(json.dumps({'error': 'Usage: transform.py <export_csv> <offer_xlsx> <template_xlsx> <output_xlsx>'}))
        sys.exit(1)
    try:
        stats = transform(sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4])
        print(json.dumps({'success': True, **stats}))
    except Exception as e:
        print(json.dumps({'error': str(e)}))
        sys.exit(1)
